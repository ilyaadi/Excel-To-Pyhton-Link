import openpyxl
import matplotlib.pyplot as plt
import numpy as np
from openpyxl.drawing.image import Image
import time

path = "C:\\Users\\aadis\\Documents\\Python\\Excel\\DragonFarmExpenses.xlsx"

workbook = openpyxl.load_workbook(path)

# print(workbook.sheetnames)

sheet = workbook['Sheet 1 - Expenses'] 
# cell_value = sheet['F1'].value 
# print(cell_value)

# cell_value = sheet.cell(row = 1, column = 1).value
# print(cell_value)


pillar_asset = 0
pillar_expense = 0

plantation_expense = 0
plantation_asset = 0

waterTank_expense = 0
waterTank_asset = 0

cowShed_expense = 0
cowShed_asset = 0

plantsMaintenance_expense = 0
plantsMaintenance_asset = 0

farmMaintenance_expense = 0
farmMaintenance_asset = 0

pillarRing_expense = 0
pillarRing_asset = 0


for p in range(90):
    desc_value = sheet.cell(row = p+1, column = 1).value
    price_value = sheet.cell(row = p+1, column = 3).value
    cat_value = sheet.cell(row = p+1, column = 5).value
    if desc_value == 'Pillar':
        # print('works')
        if cat_value == 'Asset':
            pillar_asset = pillar_asset + price_value
        elif cat_value == 'Expense':    
            pillar_expense = pillar_expense + price_value
    elif desc_value == 'Plantation':
        if cat_value == 'Asset':
            plantation_asset = plantation_asset + price_value
        elif cat_value == 'Expense':    
            plantation_expense = plantation_expense + price_value
    elif desc_value == 'Water tank':
        if cat_value == 'Asset':
            waterTank_asset = waterTank_asset + price_value
        elif cat_value == 'Expense':    
            waterTank_expense = waterTank_expense + price_value
    elif desc_value == 'Cow shed':
        if cat_value == 'Asset':
            cowShed_asset = cowShed_asset + price_value
        elif cat_value == 'Expense':    
            cowShed_expense = cowShed_expense + price_value
    elif desc_value == 'Plants Maintenance':
        if cat_value == 'Asset':
            plantsMaintenance_asset = plantsMaintenance_asset + price_value
        elif cat_value == 'Expense':    
            plantsMaintenance_expense = plantsMaintenance_expense + price_value
    elif desc_value == 'Farm Maintenance':
        if cat_value == 'Asset':
            farmMaintenance_asset = farmMaintenance_asset + price_value
        elif cat_value == 'Expense':    
            farmMaintenance_expense = farmMaintenance_expense + price_value
    elif desc_value == 'Pillar ring':
        if cat_value == 'Asset':
            pillarRing_asset = pillarRing_asset + price_value
        elif cat_value == 'Expense':    
            pillarRing_expense = pillarRing_expense + price_value

print('Pillar Asset: ', pillar_asset)
print('Pillar Expense: ', pillar_expense)
print('Plantation Asset: ', plantation_asset)
print('Plantation Expense: ', plantation_expense)
print('Water Tank Asset: ', waterTank_asset)
print('Water Tank Expense: ', waterTank_expense)  
print('Cow Shed Asset: ', cowShed_asset)
print('Cow Shed Expense: ', cowShed_expense)
print('Plants Maintenance Asset: ', plantsMaintenance_asset)
print('Plants Maintenance Expense: ', plantsMaintenance_expense)
print('Farm Maintenance Asset: ', farmMaintenance_asset)
print('Farm Maintenance Expense: ', farmMaintenance_expense)
print('Pillar Ring Asset: ', pillarRing_asset)
print('Pillar Ring Expense: ', pillarRing_expense)

total_asset = pillar_asset + plantation_asset + waterTank_asset + cowShed_asset + plantsMaintenance_asset + farmMaintenance_asset + pillarRing_asset
total_expense = pillar_expense + plantation_expense + waterTank_expense + cowShed_expense + plantsMaintenance_expense + farmMaintenance_expense + pillarRing_expense

print('Total Asset: ', total_asset)
print('Total Expense: ', total_expense)

barWidth = 0.25
fig = plt.subplots(figsize =(12, 8)) 

plt.bar(np.arange(7),
          [pillar_asset, plantation_asset, waterTank_asset, cowShed_asset, plantsMaintenance_asset, farmMaintenance_asset, pillarRing_asset], 
          width = barWidth, 
          label = 'Asset', 
          color = 'blue')
plt.bar(np.arange(7) + barWidth,
          [pillar_expense, plantation_expense, waterTank_expense, cowShed_expense, plantsMaintenance_expense, farmMaintenance_expense, pillarRing_expense], 
          width = barWidth, 
          label = 'Expense', 
          color = 'orange')

plt.xlabel('Categories', fontweight ='bold', fontsize = 15)
plt.ylabel('Amount', fontweight ='bold', fontsize = 15)
plt.xticks(np.arange(7) + barWidth / 2, 
           ['Pillar', 'Plantation', 'Water Tank', 'Cow Shed', 'Plants Maintenance', 'Farm Maintenance', 'Pillar Ring'])
plt.title('Dragon Farm Expenses', fontweight ='bold', fontsize = 20)
plt.legend()
plt.grid(axis='y', linestyle='--', alpha=0.7)
plt.tight_layout()

plt.pause(20)  


# plt.savefig('images/DragonFarmExpenses.png', dpi=300, bbox_inches='tight')
# img = openpyxl.drawing.image.Image('images/DragonFarmExpenses.png')

plt.savefig('C:\\Users\\aadis\\Documents\\Python\\Excel\\DragonFarmExpenses.png', dpi=300, bbox_inches='tight')
img = openpyxl.drawing.image.Image('C:\\Users\\aadis\\Documents\\Python\\Excel\\DragonFarmExpenses.png')
plt.close()

img.anchor = 'H1'  
sheet.add_image(img)

workbook.save("C:/Users/aadis/Documents/Python/Excel/DragonFarmExpensesUpdated.xlsx")

